# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
import arrow
from scrapy.mail import MailSender
import scrapy.settings.default_settings
import me.settings
from highcharts import Highchart
import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
from email.mime.application import MIMEApplication  
import os
# openpyxl库参考文档：http://openpyxl.readthedocs.org/en/2.3.3/
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import Style,Font,Alignment,Fill,Color,PatternFill,Border,Side


class dataPipeline(object):
	def __init__(self):
		print "Scrapying..."
		self._week=arrow.now().span("week")#取某天所在年的端点
		#self._week=arrow.get("2016-03-06").span("week")#统计区间单位为week
		self.start=self._week[0] #起点
		self.end=self._week[1]	#终点

		self.week_text=u"{0}至{1}\n".format(self.start.format("YYYY-MM-DD"),self.end.format("YYYY-MM-DD"))
		self.head=u"<div style='border-bottom: 1px solid #dddddd;margin-bottom:10px;'>老师好!本周学院网站更新了如下文章:</div>".encode("gbk")
		self.body=""

		self.body_end=u"""
		<div style='border-top: 1px solid #dddddd;margin-top:20px;margin-bottom:100px'>
			统计信息见附件(<span style='color:#ff0000'>注意事项：如果打开Excel出现合计部分不显示，是因为公式被屏蔽，请点击上方弹出的按钮【启用编辑】</span >)
		</div>
		<div>Me工作室 {0}</div>
		<div>本邮件为系统自动发送，如有问题请联系hanxiaomax@qq.com</div>
		""".format(arrow.now()).encode('gbk')
		

		#是否在生产服务器上，本机调试时请设置settings.py中的IS_PRODUCTION=False
		if me.settings.IS_PRODUCTION:
			self.recipients=[
								"yifan.zhou@seu.edu.cn",
								"lingfeng_ai@foxmail.com",
								"1205578775@qq.com" #尹
							]
		else:
			self.recipients=["1196529490@qq.com"]
		#需要统计的内容
		#添加新的部门请插入到【其他】前面
		#append new sector in the last 2nd pos
		#需要删除原来的文件才能更新（请备份数据）
		self.summary=[
		{"sector":u"教师事务中心","sum":0,"publishers":dict.fromkeys((u"汤蓓",u"陈科",u"闫焱",u"贾方",u"张琪",u"方春富",u"李天"),0)},
		{"sector":u"本科生事务中心","sum":0,"publishers":dict.fromkeys((u"李晓燕",u"金传志",u"殷超",u"梅仂盈",u"史红叶",u"刘宗涛"),0)},
		{"sector":u"研究生事务中心","sum":0,"publishers":dict.fromkeys((u"陈斌",u"夏敏思",u"周双",u"徐志芳",u"陈雨",u"王赛君",u"唐攀",u"冒明山"),0)},
		{"sector":u"学院党政领导","sum":0,"publishers":dict.fromkeys((u"王斌",u"殷国栋",u"孙蓓蓓",u"陈云飞",u"周一帆",u"黄鹏",u"周怡君",u"沙菁洁"),0)},
		{"sector":u"ME工作室","sum":0,"publishers":dict.fromkeys((u"谭会明",u"倪光一",u"吴宣勇",u"柏硕",u"贺从愿",u"尹德军",u"艾凌风"),0)},
		# {"sector":u"其他","sum":0,"publishers":{}}
		]
		

	def process_item(self, item, spider):
		"处理每个通过管线的数据包，在这里就是采集到一条标题"
		time=arrow.get(item["time"],"YYYY-MM-DD")
		title=item["title"]
		publisher=item["publisher"]
		#day表示，把start和end区间内的日期拆分成以day为单位
		#取.date()可以忽略时间
		week=[i.date() for i in arrow.Arrow.range('day', self.start, self.end)]
		if time.date() in week:
			text="{0}\t{1}\t{2}\n".format(title,time.date(),publisher)#构造正文
			self.body+=("<ul style='padding:5px; color:#228B22'>%s</ul>"%(text))
			self.caculate(publisher)


		
	def caculate(self,publisher):
		"计算某个人发文数"
		try:
			for sector in self.summary:
				if sector["publishers"].has_key(publisher.decode('gbk')):
					sector["publishers"][publisher.decode('gbk')]+=1
		except:#此人不在统计范围内
				pass

	def close_spider(self,spider):
		"在关闭爬虫前进行excel生成，同时发送邮件"
		filename="./{0}-{1}".format(self.start.format("YYYY-MM-DD"),self.end.format("YYYY-MM-DD"))#文件名，不包含扩展名
		print "closing..."
		print self.start,self.end
		try:
			#备份数据
			os.system("cp {0}.xlsx {1}.xlsx".format(u"summary".encode('gbk'),self.end.format("YYYY-MM-DD")))
		except Exception,e:
			print e
		maker = ExcelManipulation(self.summary,u"{0}至{1}".format(self.start.format("YYYY-MM-DD"),self.end.format("YYYY-MM-DD")))
		maker.run()
		maker.saveAs(u"summary")#保存excel


		email_message=self.head+(self.body if self.body is not "" else u"无".encode('gbk'))+self.body_end
		with open('artical.txt','w+') as f:
			f.write(email_message)
		self._send_mail(filename,email_message,u"学院网站每周总结({0})".format(self.week_text),self.recipients)#发送邮件
		

	@staticmethod
	def _send_mail(filename,message,title,recipients):
		"""
		发送电子邮件
		@filename:附件文件名
		@message：邮件正文
		@title：邮件主题
		@recipients:收件人列表
		"""
		print "Sending mail..........."
		#只能用学校的邮箱，注意密码安全。最好放到环境变量里面
		mailUser = '220150233@seu.edu.cn'
		mailPassword ='wqy1177393'
		msg = MIMEMultipart()
		msg['From'] = mailUser
		msg['To'] = ",".join(recipients)
		msg['Subject'] = title
		
		xls_part = MIMEApplication(open("summary.xlsx",'rb').read()) 


		# chart_part = MIMEApplication(open(filename+".html",'rb').read()) 
		xls_part.add_header('Content-Disposition', 'attachment', filename="summary.xlsx")  
		# chart_part.add_header('Content-Disposition', 'attachment', filename=filename+".html")  
		
		msg.attach(MIMEText(message, 'html', 'gbk')) #正文
		msg.attach(xls_part)  #excel附件
		# msg.attach(chart_part) #表格html附件

		try:
			mailServer = smtplib.SMTP('smtp.seu.edu.cn', 25)
			mailServer.ehlo()
			mailServer.starttls()
			mailServer.ehlo()
			mailServer.login(mailUser, mailPassword)
			mailServer.sendmail(mailUser, recipients, msg.as_string())
			mailServer.close()
			print "Mail sent!"
		except Exception, e:
			print "Failed!"
			print e
			raise e
			

	@staticmethod
	def _saveChart(summary,filename):
		"生成图表，附加功能，并没有投入使用"
		chart = Highchart(width=850, height=400)
		options = {'chart': {
        'type': 'column'
	    },'title': {
	        'text': u'机械学院网站发文统计'
	    },'subtitle': {
	        'text': u'点击柱状图查看各事务中心明细'
	    },'xAxis': {
	        'type': 'category'
	    },'yAxis': {
	        'title': {
	            'text': u'发文数',
	        },
	        "minTickInterval":1,
	        "allowDecimals":"true"
	    },'legend': {
	        'enabled': False
	    },'plotOptions': {
	        'series': {
	            'borderWidth': 0,
	            'dataLabels': {
	                'enabled': True,
	                'format': '{point.y}篇'
	            }
	        }
	    },'tooltip': {
	        'headerFormat': '<span style="font-size:11px">{series.name}</span><br>',
	        'pointFormat': '<span style="color:{point.color}">{point.name}</span>: <b>{point.y}篇</b><br/>'
	    }, 
		}
		data=[dict(name=i["sector"],y=i["sum"],drilldown=i["sector"]) for i in summary]#分离合计数据
		chart.set_dict_options(options)
		chart.add_data_set(data,'column', "事务中心", colorByPoint= True)
		for i in range(len(summary)):#分离并设置publisher数据
			drill_data=[[p,num] for p,num in summary[i]["publishers"].items()]
			chart.add_drilldown_data_set(drill_data, 'column', summary[i]["sector"], name=summary[i]["sector"] )

		chart.save_file(filename)
	

class ExcelManipulation(object):
	"这个类负责整个excel的操作"
	def __init__(self,data,interval):
		self.data=data
		self.interval=interval
		try:#尝试打开原有的excel
			self.wb=load_workbook(u"summary.xlsx")
			self.sheet=self.wb.active
			self.firstTime=False
		except:#新建excel
			print "Creating new excel..."
			self.wb = Workbook()
			self.sheet=self.wb.active
			self.sheet['A2']=u"区间"
			self.sheet.column_dimensions["A"].width=23
			start=2
			for index,item in enumerate(data):
				end=start+len(item["publishers"])
				self.sheet.merge_cells(start_row=1,start_column=start,end_row=1,end_column=end)#合并单元格
				_cell = self.sheet.cell(row=1,column=start)
				_cell.value = item["sector"]#写单元格(一个部门)
				_style=self.getStyle(index)#根据获取样式
				_cell.font = _style['font']
				_cell.alignment=_style['alignment']
				_cell.fill=_style['fill']
				sumCell=self.sheet.cell(row=2,column=end)
				sumCell.value=u"本周合计"
				sumCell.font = _style['font']
				sumCell.alignment=_style['alignment']
				sumCell.fill=_style['fill']
				self.sheet.column_dimensions[get_column_letter(end)].width=8#设置单元格宽度
				start=end+1
			self.firstTime=True
		
	@staticmethod
	def getStyle(index):
		"获取一个不同的颜色风格"
		colorMap=["99CCFF","CCFFCC","FF9900","FFFFE0","48D1CC","FFB6C1","90EE90"]
		if index >= len(colorMap):
			hex_color=colorMap[index-len(colorMap) if index >= len(colorMap) else colorMap[index]]
		else:
			hex_color=colorMap[index]
		font = Font(name='Microsoft Yahei',size=11)
		alignment=Alignment(horizontal='center',vertical='bottom')
		fill = PatternFill(start_color=Color(rgb=hex_color),end_color=Color(rgb=hex_color),fill_type='solid')
		return {'font':font,"alignment":alignment,"fill":fill}
		
	def run(self):
		"生成excel"
		pos=1#column position
		rowCount=len(self.sheet.rows)#计算当前总行数
		#如果是第一次生成，这里应该是2行
		if self.firstTime:
			self.start_row=rowCount+1
		else:
			self.start_row=rowCount#需要覆盖最后一行

		
		intervalCell=self.sheet.cell(row = self.start_row, column = pos)
		intervalCell.value=self.interval
		#####################为本周的数据创建一行内容，包括每行的合计############################
		for i in range(len(self.data)):
			for (name,num)in self.data[i]["publishers"].items():
				pos+=1
				if self.firstTime:#首次创建文件
					nameCell=self.sheet.cell(row = self.start_row-1, column = pos)
					nameCell.value=name
					#起始行+1
				numCell=self.sheet.cell(row = self.start_row, column = pos)
				numCell.value=num
				numCell.fill=PatternFill(start_color=Color(rgb="ffffff"),end_color=Color(rgb="ffffff"),fill_type='solid')
				numCell.border=self.makeBorder()

				self.sheet.column_dimensions[get_column_letter(pos)].width=7
			pos+=1
			#为每个部门生成该周合计
			sumCell=self.sheet.cell(row = self.start_row, column = pos)
			sumCell.fill=PatternFill(start_color=Color(rgb="ffffff"),end_color=Color(rgb="ffffff"),fill_type='solid')
			sumCell.value="=SUM({0}{2}:{1}{2})".format(
													get_column_letter(pos-len(self.data[i]["publishers"])),#起点
													get_column_letter(pos-1),#终点letter
													self.start_row)
			sumCell.border=self.makeBorder()
		####################总计#############################
		rowCount=len(self.sheet.rows)#再次计算行数
		self.sheet["A%d" %(rowCount+1)]=u"总计"
		col=1
		for i in range(len(self.data)):#每个机构
			for j in range(len(self.data[i]["publishers"])):#机构中的每个人
				col+=1
				totalCell=self.sheet.cell(row = rowCount+1,column = col)
				totalCell.value="=SUM({0}3:{0}{1})".format(get_column_letter(col),rowCount)#每个人的总计
				totalCell.fill=PatternFill(start_color=Color(rgb="DCDCDC"),end_color=Color(rgb="DCDCDC"),fill_type='solid')
			col+=1
			totalCell=self.sheet.cell(row = rowCount+1,column = col)
			totalCell.value="=SUM({0}3:{0}{1})".format(get_column_letter(col),rowCount)#合计的总计
			totalCell.fill=PatternFill(start_color=Color(rgb="DCDCDC"),end_color=Color(rgb="DCDCDC"),fill_type='solid')
	
		
		
	
	def makeBorder(self,style="thin",color="D4D4D4"):
		"返回指定的一个Border样式"
		return Border(left=Side(border_style=style,color=color),
				right=Side(border_style=style,color=color),
                top=Side(border_style=style,color=color),
                bottom=Side(border_style=style,color=color),
                outline=Side(border_style=style,color=color),
                vertical=Side(border_style=style,color=color),
                horizontal=Side(border_style=style,color=color)
               )

	def saveAs(self,filename):
		print "Excel Saving..."
		try:
			self.wb.save(filename+".xlsx")
			print "Done!"
		except Exception,e:
			print e
			raise Exception
			